from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from django.db import transaction
from .models import Order, OrderItem
from .serializers import OrderSerializer, OrderCreateSerializer
from cart.models import Cart
from products.models import Product, ProductSize
from alerts.services import create_stock_alert

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_order(request):
    """Create order from cart"""
    serializer = OrderCreateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    from django.conf import settings
    admin_email = request.headers.get('X-Admin-Email', getattr(settings, 'DEFAULT_FROM_EMAIL', 'm.mesum1800@gmail.com'))
    admin_phone = request.headers.get('X-Admin-Phone', getattr(settings, 'ADMIN_WHATSAPP_NUMBER', ''))

    # Fix for placeholder email causing bounces
    if not admin_email or 'libaassapna.com' in admin_email:
        from django.conf import settings
        admin_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'm.mesum1800@gmail.com')

    cart_items = Cart.objects.filter(user=request.user)
    if not cart_items.exists():
        return Response(
            {'error': 'Cart is empty'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Calculate total
    total = sum(item.subtotal for item in cart_items)

    # Check stock availability
    for cart_item in cart_items:
        product = cart_item.product
        available_stock = product.stock_qty
        if cart_item.size:
            try:
                ps = ProductSize.objects.get(product=product, size=cart_item.size)
                available_stock = ps.stock_qty
            except ProductSize.DoesNotExist:
                pass

        # Unstitched products: 1 suit = 4 meters of fabric
        required_stock = cart_item.quantity * Product.METERS_PER_SUIT if product.is_unstitched else cart_item.quantity

        if available_stock < required_stock:
            size_info = f" (Size: {cart_item.size})" if cart_item.size else ""
            if product.is_unstitched:
                suits_available = available_stock // Product.METERS_PER_SUIT
                return Response(
                    {'error': f'Insufficient fabric for {product.name}. Only {suits_available} suit(s) worth of fabric available ({available_stock} meters).'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            return Response(
                {'error': f'Insufficient stock for {product.name}{size_info}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    with transaction.atomic():
        # Create order
        order = Order.objects.create(
            customer=request.user,
            total_amount=total,
            payment_method=serializer.validated_data['payment_method'],
            payment_status=serializer.validated_data.get('payment_status', 'pending'),
            order_status=serializer.validated_data.get('order_status', 'pending_payment'),
            shipping_address=serializer.validated_data['shipping_address'],
            shipping_city=serializer.validated_data['shipping_city'],
            shipping_phone=serializer.validated_data['shipping_phone'],
            transaction_id=serializer.validated_data.get('transaction_id', ''),
        )

        # Create order items and update stock
        for cart_item in cart_items:
            OrderItem.objects.create(
                order=order,
                product=cart_item.product,
                quantity=cart_item.quantity,
                price=cart_item.product.final_price,
                size=cart_item.size
            )

            # # Update product stock
            # cart_item.product.stock_qty -= cart_item.quantity
            # cart_item.product.save()
            
            # # LOW STOCK CHECK
            # if cart_item.product.stock_qty <= cart_item.product.low_stock_threshold and cart_item.product.stock_qty > 0:
            #     create_stock_alert(cart_item.product, 'low')

            # # OUT OF STOCK CHECK
            # elif cart_item.product.stock_qty <= 0:
            #     create_stock_alert(cart_item.product, 'out')
            # Update stock
            product = cart_item.product
            size_str = cart_item.size

            # Unstitched: deduct meters_per_suit * quantity; stitched: deduct quantity
            deduction = cart_item.quantity * Product.METERS_PER_SUIT if product.is_unstitched else cart_item.quantity
            
            if size_str:
                try:
                    product_size = ProductSize.objects.get(product=product, size=size_str)
                    product_size.stock_qty -= deduction
                    if product_size.stock_qty < 0:
                        product_size.stock_qty = 0
                    product_size.save()
                    # Product.stock_qty is synced via signal
                except ProductSize.DoesNotExist:
                    product.stock_qty -= deduction
                    if product.stock_qty < 0:
                        product.stock_qty = 0
                    product.save()
            else:
                product.stock_qty -= deduction
                if product.stock_qty < 0:
                    product.stock_qty = 0
                product.save()

            # Reload fresh value (IMPORTANT FIX)
            product.refresh_from_db()
            # ALERT LOGIC (SINGLE SOURCE OF TRUTH)
            # 1. Total Product Alert
            if product.stock_qty <= 0:
                create_stock_alert(product, 'out', admin_email=admin_email, admin_phone=admin_phone)
            elif product.stock_qty <= product.low_stock_threshold:
                create_stock_alert(product, 'low', admin_email=admin_email, admin_phone=admin_phone)

            # 2. Specific Size Alert
            if size_str:
                try:
                    ps = ProductSize.objects.get(product=product, size=size_str)
                    if ps.stock_qty <= 0:
                        create_stock_alert(product, 'out', size=size_str, admin_email=admin_email, admin_phone=admin_phone)
                    elif ps.stock_qty <= product.low_stock_threshold:
                        create_stock_alert(product, 'low', size=size_str, admin_email=admin_email, admin_phone=admin_phone)
                except ProductSize.DoesNotExist:
                    pass
        # Clear cart
        cart_items.delete()

        # Record sales analytics
        try:
            from analytics.models import SalesAnalytics
            for item in order.items.all():
                SalesAnalytics.objects.create(
                    product=item.product,
                    order=order,
                    quantity_sold=item.quantity,
                    revenue=item.subtotal,
                    sale_date=order.order_date.date()
                )
        except:
            pass  # Analytics app might not be set up yet

    # Generate payment URL if needed
    payment_url = None
    if serializer.validated_data['payment_method'] in ['easypaisa', 'jazzcash']:
        try:
            from payments.services import get_payment_gateway
            gateway = get_payment_gateway(serializer.validated_data['payment_method'])
            payment_data = gateway.create_payment(order.order_id, total)
            payment_url = payment_data.get('payment_url')
        except:
            pass

    # Send notifications (email + WhatsApp) asynchronously
    import threading
    
    def send_notifications_bg(order_obj, admin_email_str, admin_phone_str):
        try:
            from .notifications import (
                send_order_notification_email,
                send_customer_order_confirmation_email,
                send_order_whatsapp,
            )
            # 1) Order placed notification
            send_order_notification_email(order_obj, admin_email_str)
            send_customer_order_confirmation_email(order_obj)
            send_order_whatsapp(order_obj, admin_phone_str)
        except Exception as e:
            print(f"[NOTIFY] Background notification error: {e}")

    # Start the background thread
    threading.Thread(target=send_notifications_bg, args=(order, admin_email, admin_phone)).start()

    return Response({
        'order_id': order.order_id,
        'total_amount': float(total),
        'payment_url': payment_url,
        'message': 'Order created successfully'
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_order(request, order_id):
    """Get order details"""
    try:
        order = Order.objects.get(order_id=order_id, customer=request.user)
    except Order.DoesNotExist:
        return Response(
            {'error': 'Order not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    serializer = OrderSerializer(order)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_orders(request):
    """Get all orders for current user, or all orders for admin"""
    # Admin users can see all orders
    if request.user.is_staff or request.user.is_superuser or getattr(request.user, 'role', None) == 'admin':
        orders = Order.objects.all().order_by('-order_date')
    else:
        orders = Order.objects.filter(customer=request.user).order_by('-order_date')
    
    serializer = OrderSerializer(orders, many=True)
    return Response({
        'total': orders.count(),
        'orders': serializer.data
    })


@api_view(['PATCH'])
@permission_classes([IsAdminUser])
def update_order_status(request, order_id):
    """Update order status (Admin only)"""
    try:
        order = Order.objects.get(order_id=order_id)
    except Order.DoesNotExist:
        return Response(
            {'error': 'Order not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    new_status = request.data.get('status')
    if not new_status:
        return Response(
            {'error': 'Status is required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    valid_statuses = [choice[0] for choice in Order.ORDER_STATUS_CHOICES]
    if new_status not in valid_statuses:
        return Response(
            {'error': 'Invalid status'},
            status=status.HTTP_400_BAD_REQUEST
        )

    order.order_status = new_status
    if new_status == 'paid':
        order.payment_status = 'paid'
    order.save()

    # Notify customer about status update
    try:
        from .notifications import send_customer_status_update_notification
        send_customer_status_update_notification(order)
    except Exception as e:
        print(f"[NOTIFY] Status update notification failed: {e}")

    return Response({
        'message': 'Order status updated',
        'order_id': order.order_id,
        'status': order.order_status
    })

@api_view(['GET'])
@permission_classes([IsAdminUser])
def get_orders_by_user(request, user_id):
    """Get all orders for a specific user (Admin only)"""
    orders = Order.objects.filter(customer__user_id=user_id).order_by('-order_date')
    serializer = OrderSerializer(orders, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([])  # We handle auth manually below to support token-in-URL
def export_sales_report(request):
    """Export sales report as Excel file for a date range (Admin only)"""
    import io
    from datetime import datetime
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Authenticate via query param token (for direct browser downloads)
    user = request.user
    if not user or not user.is_authenticated:
        token_str = request.query_params.get('token')
        if token_str:
            from rest_framework_simplejwt.tokens import AccessToken
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                access_token = AccessToken(token_str)
                user = User.objects.get(pk=access_token['user_id'])
            except Exception:
                return Response({'error': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

    # Check admin permission
    if not (user.is_staff or user.is_superuser or getattr(user, 'role', None) == 'admin'):
        return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')

    if not date_from or not date_to:
        return Response(
            {'error': 'Both date_from and date_to are required (YYYY-MM-DD)'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        start_date = datetime.strptime(date_from, '%Y-%m-%d')
        end_date = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except ValueError:
        return Response(
            {'error': 'Invalid date format. Use YYYY-MM-DD'},
            status=status.HTTP_400_BAD_REQUEST
        )

    orders = Order.objects.filter(
        order_date__gte=start_date,
        order_date__lte=end_date
    ).exclude(order_status='cart').prefetch_related('items__product', 'customer').order_by('-order_date')

    # ── Create Workbook ──────────────────────────────────────────────────────
    wb = Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=16, color='0F172A')
    subtitle_font = Font(name='Calibri', size=11, color='64748B')
    money_font = Font(name='Calibri', bold=True, size=11, color='16A34A')
    thin_border = Border(
        bottom=Side(style='thin', color='E2E8F0')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(wrap_text=True, vertical='center')

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Orders Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws_orders = wb.active
    ws_orders.title = 'Orders Summary'

    # Title rows
    ws_orders.merge_cells('A1:I1')
    ws_orders['A1'] = 'LIBAAS SAPNA — Sales Report'
    ws_orders['A1'].font = title_font
    ws_orders.merge_cells('A2:I2')
    ws_orders['A2'] = f'Period: {date_from} to {date_to}'
    ws_orders['A2'].font = subtitle_font

    # Header row
    headers = [
        'Order #', 'Date', 'Customer', 'Email', 'City',
        'Items', 'Payment Method', 'Status', 'Total (Rs.)'
    ]
    col_widths = [10, 18, 22, 28, 16, 8, 18, 14, 16]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws_orders.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws_orders.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    row = 5
    total_revenue = 0
    status_counts = {}
    payment_counts = {}

    for order in orders:
        item_count = order.items.count()
        customer_name = order.customer.name if hasattr(order.customer, 'name') else str(order.customer)
        customer_email = order.customer.email if hasattr(order.customer, 'email') else ''

        ws_orders.cell(row=row, column=1, value=order.order_id).alignment = center_align
        ws_orders.cell(row=row, column=2, value=order.order_date.strftime('%Y-%m-%d %H:%M'))
        ws_orders.cell(row=row, column=3, value=customer_name)
        ws_orders.cell(row=row, column=4, value=customer_email)
        ws_orders.cell(row=row, column=5, value=order.shipping_city)
        ws_orders.cell(row=row, column=6, value=item_count).alignment = center_align
        ws_orders.cell(row=row, column=7, value=order.get_payment_method_display()).alignment = center_align
        ws_orders.cell(row=row, column=8, value=order.get_order_status_display()).alignment = center_align
        amount_cell = ws_orders.cell(row=row, column=9, value=float(order.total_amount))
        amount_cell.number_format = '#,##0.00'
        amount_cell.alignment = Alignment(horizontal='right', vertical='center')

        for c in range(1, 10):
            ws_orders.cell(row=row, column=c).border = thin_border

        total_revenue += float(order.total_amount)
        status_counts[order.get_order_status_display()] = status_counts.get(order.get_order_status_display(), 0) + 1
        payment_counts[order.get_payment_method_display()] = payment_counts.get(order.get_payment_method_display(), 0) + 1
        row += 1

    # Total row
    row += 1
    ws_orders.cell(row=row, column=8, value='TOTAL REVENUE:').font = Font(bold=True, size=12)
    total_cell = ws_orders.cell(row=row, column=9, value=total_revenue)
    total_cell.font = Font(bold=True, size=12, color='16A34A')
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal='right')

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Item Details
    # ══════════════════════════════════════════════════════════════════════════
    ws_items = wb.create_sheet('Item Details')

    ws_items.merge_cells('A1:H1')
    ws_items['A1'] = 'Order Items Breakdown'
    ws_items['A1'].font = title_font

    item_headers = [
        'Order #', 'Product', 'Size', 'Qty', 'Unit Price (Rs.)',
        'Subtotal (Rs.)', 'Category', 'Customer'
    ]
    item_widths = [10, 30, 10, 8, 16, 16, 16, 22]
    for col_idx, (header, width) in enumerate(zip(item_headers, item_widths), 1):
        cell = ws_items.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws_items.column_dimensions[get_column_letter(col_idx)].width = width

    item_row = 4
    for order in orders:
        customer_name = order.customer.name if hasattr(order.customer, 'name') else str(order.customer)
        for item in order.items.all():
            product_name = item.product.name if item.product else 'Deleted Product'
            category = ''
            if item.product and hasattr(item.product, 'category'):
                cat = item.product.category
                category = cat.name if hasattr(cat, 'name') else str(cat) if cat else ''

            ws_items.cell(row=item_row, column=1, value=order.order_id).alignment = center_align
            ws_items.cell(row=item_row, column=2, value=product_name)
            ws_items.cell(row=item_row, column=3, value=item.size or '—').alignment = center_align
            ws_items.cell(row=item_row, column=4, value=item.quantity).alignment = center_align
            price_cell = ws_items.cell(row=item_row, column=5, value=float(item.price))
            price_cell.number_format = '#,##0.00'
            sub_cell = ws_items.cell(row=item_row, column=6, value=float(item.subtotal))
            sub_cell.number_format = '#,##0.00'
            ws_items.cell(row=item_row, column=7, value=category).alignment = center_align
            ws_items.cell(row=item_row, column=8, value=customer_name)

            for c in range(1, 9):
                ws_items.cell(row=item_row, column=c).border = thin_border
            item_row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws_summary = wb.create_sheet('Summary')

    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = 'Sales Summary'
    ws_summary['A1'].font = title_font

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 5
    ws_summary.column_dimensions['D'].width = 25
    ws_summary.column_dimensions['E'].width = 20

    # Key Metrics
    summary_header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    ws_summary.cell(row=3, column=1, value='KEY METRICS').font = Font(bold=True, size=12)

    metrics = [
        ('Total Orders', len(orders)),
        ('Total Revenue', f'Rs. {total_revenue:,.2f}'),
        ('Average Order Value', f'Rs. {(total_revenue / len(orders)):,.2f}' if orders else 'Rs. 0'),
        ('Report Period', f'{date_from} to {date_to}'),
    ]
    for i, (label, value) in enumerate(metrics, 4):
        ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=str(value))

    # Order Status Breakdown
    ws_summary.cell(row=10, column=1, value='BY ORDER STATUS').font = Font(bold=True, size=12)
    r = 11
    for s, count in sorted(status_counts.items(), key=lambda x: -x[1]):
        ws_summary.cell(row=r, column=1, value=s)
        ws_summary.cell(row=r, column=2, value=count)
        r += 1

    # Payment Method Breakdown
    ws_summary.cell(row=10, column=4, value='BY PAYMENT METHOD').font = Font(bold=True, size=12)
    r = 11
    for pm, count in sorted(payment_counts.items(), key=lambda x: -x[1]):
        ws_summary.cell(row=r, column=4, value=pm)
        ws_summary.cell(row=r, column=5, value=count)
        r += 1

    # ── Write to response ─────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'LibaasSapna_SalesReport_{date_from}_to_{date_to}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.document'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

